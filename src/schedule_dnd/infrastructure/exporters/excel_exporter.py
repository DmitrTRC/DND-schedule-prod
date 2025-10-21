"""
Excel export implementation.

Author: DmitrTRC
"""

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schedule_dnd.domain.exceptions import ExportFailedError
from schedule_dnd.domain.models import Schedule
from schedule_dnd.infrastructure.config.settings import get_settings
from schedule_dnd.infrastructure.exporters.base import BaseExporter


class ExcelExporter(BaseExporter):
    """Exporter for Excel (XLSX) format."""

    def __init__(self) -> None:
        """Initialize Excel exporter."""
        self.settings = get_settings()

    def export(self, schedule: Schedule, output_path: Optional[Path] = None) -> Path:
        """
        Export schedule to Excel file.

        Args:
            schedule: Schedule to export
            output_path: Optional output file path

        Returns:
            Path to the exported file

        Raises:
            ExportError: If export fails
        """
        self.validate_schedule(schedule)

        if output_path is None:
            filename = self.get_default_filename(schedule)
            output_path = self.settings.output_dir / filename

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "График дежурств"

            # Add title
            month_name = schedule.metadata.month.display_name()
            title = f"График дежурств ДНД - {month_name} {schedule.metadata.year}"
            ws.append([title])

            # Style title
            title_cell = ws["A1"]
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal="center")
            ws.merge_cells("A1:F1")

            # Add empty row
            ws.append([])

            # Add headers
            headers = [
                "Подразделение",
                "Дата",
                "День недели",
                "Тип дежурства",
                "Время",
                "Примечания",
            ]
            ws.append(headers)

            # Style headers
            header_row = ws[3]
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)

            for cell in header_row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add data
            for unit in schedule.units:
                for shift in unit.shifts:
                    ws.append(
                        [
                            unit.unit_name,
                            shift.date,
                            shift.get_day_of_week(),
                            shift.duty_type.value,
                            shift.time,
                            shift.notes,
                        ]
                    )

            # Apply borders to all cells with data
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for row in ws.iter_rows(
                min_row=3, max_row=ws.max_row, min_col=1, max_col=6
            ):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Auto-adjust column widths
            column_widths = {
                "A": 35,  # Подразделение
                "B": 12,  # Дата
                "C": 15,  # День недели
                "D": 15,  # Тип дежурства
                "E": 15,  # Время
                "F": 50,  # Примечания
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Set row height for title
            ws.row_dimensions[1].height = 25

            # Add metadata at the bottom if enabled
            if self.settings.include_metadata:
                ws.append([])
                ws.append([])
                ws.append(["Источник:", schedule.metadata.source or ""])
                ws.append(["Подписант:", schedule.metadata.signatory or ""])

                if schedule.metadata.note:
                    ws.append([])
                    ws.append(["Примечание:"])
                    ws.append([schedule.metadata.note])

            # Set document properties
            wb.properties.creator = self.settings.excel_author
            wb.properties.title = title

            # Save workbook
            wb.save(output_path)

            return output_path

        except Exception as e:
            raise ExportFailedError(
                format_type="Excel", reason=f"Failed to create Excel file: {str(e)}"
            ) from e

    def get_file_extension(self) -> str:
        """Get file extension for Excel."""
        return "xlsx"

    def get_format_name(self) -> str:
        """Get format name."""
        return "Excel"
