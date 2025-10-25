"""
Comprehensive unit tests for all exporters

This module contains comprehensive tests for:
- BaseExporter (abstract functionality)
- CSVExporter
- JSONExporter
- ExcelExporter
- MarkdownExporter
- HTMLExporter
- ExporterFactory

Author: DmitrTRC
"""

import csv
import json
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from schedule_dnd.domain.enums import DutyType, ExportFormat, Month
from schedule_dnd.domain.exceptions import ExportFailedError, ValidationError
from schedule_dnd.domain.models import Schedule, ScheduleMetadata, Shift, Unit
from schedule_dnd.infrastructure.exporters.base import BaseExporter
from schedule_dnd.infrastructure.exporters.csv_exporter import CSVExporter
from schedule_dnd.infrastructure.exporters.excel_exporter import ExcelExporter
from schedule_dnd.infrastructure.exporters.factory import ExporterFactory
from schedule_dnd.infrastructure.exporters.html_exporter import HTMLExporter
from schedule_dnd.infrastructure.exporters.json_exporter import JSONExporter
from schedule_dnd.infrastructure.exporters.markdown_exporter import MarkdownExporter

# ========== Fixtures ==========


@pytest.fixture
def sample_schedule():
    """Create a sample schedule for testing."""
    metadata = ScheduleMetadata(
        document_type="График дежурств",
        month=Month.OCTOBER,
        year=2025,
        created_at=datetime(2025, 10, 1, 10, 0, 0),
        created_by="Test User",
        source="Test Source",
        signatory="Test Signatory",
        note="Test note",
    )

    unit1 = Unit(
        id=1,
        unit_name="ДНД «Всеволожский дозор»",
        shifts=[
            Shift(
                date="01.10.2025",
                duty_type=DutyType.PDN,
                time="09:00-18:00",
                notes="Утреннее дежурство",
            ),
            Shift(
                date="02.10.2025",
                duty_type=DutyType.PPSP,
                time="18:00-22:00",
                notes="Ночное дежурство",
            ),
        ],
    )

    unit2 = Unit(
        id=2,
        unit_name="ДНД «Заневское ГП»",
        shifts=[
            Shift(
                date="03.10.2025",
                duty_type=DutyType.UUP,
                time="09:00-18:00",
                notes="Дневное дежурство",
            ),
        ],
    )

    return Schedule(metadata=metadata, units=[unit1, unit2])


@pytest.fixture
def minimal_schedule():
    """Create a minimal schedule with one unit and one shift."""
    metadata = ScheduleMetadata(
        document_type="График дежурств",
        month=Month.NOVEMBER,
        year=2025,
        created_at=datetime(2025, 11, 1, 10, 0, 0),
        created_by="Test User",
    )

    unit = Unit(
        id=1,
        unit_name="ДНД «Всеволожский дозор»",
        shifts=[
            Shift(
                date="01.11.2025",
                duty_type=DutyType.PDN,
                time="09:00-18:00",
                notes="",
            ),
        ],
    )

    return Schedule(metadata=metadata, units=[unit])


@pytest.fixture
def empty_schedule():
    """Create a schedule with no units (invalid)."""
    metadata = ScheduleMetadata(
        document_type="График дежурств",
        month=Month.DECEMBER,
        year=2025,
        created_at=datetime(2025, 12, 1, 10, 0, 0),
        created_by="Test User",
    )

    return Schedule(metadata=metadata, units=[])


@pytest.fixture
def temp_output_dir(tmp_path):
    """Create a temporary output directory."""
    output_dir = tmp_path / "exports"
    output_dir.mkdir(exist_ok=True)
    return output_dir


# ========== Base Exporter Tests ==========


class TestBaseExporter:
    """Tests for BaseExporter abstract class."""

    def test_get_default_filename(self, sample_schedule):
        """Test default filename generation."""

        class TestExporter(BaseExporter):
            def export(self, schedule, output_path=None):
                pass

            def get_file_extension(self):
                return "test"

            def get_format_name(self):
                return "Test"

        exporter = TestExporter()
        filename = exporter.get_default_filename(sample_schedule)

        assert filename == "schedule_2025_10.test"
        assert "2025" in filename
        assert "10" in filename
        assert filename.endswith(".test")

    def test_validate_schedule_success(self, sample_schedule):
        """Test successful schedule validation."""

        class TestExporter(BaseExporter):
            def export(self, schedule, output_path=None):
                pass

            def get_file_extension(self):
                return "test"

            def get_format_name(self):
                return "Test"

        exporter = TestExporter()
        assert exporter.validate_schedule(sample_schedule) is True

    def test_validate_schedule_fails_empty(self, empty_schedule):
        """Test validation fails for empty schedule."""

        class TestExporter(BaseExporter):
            def export(self, schedule, output_path=None):
                pass

            def get_file_extension(self):
                return "test"

            def get_format_name(self):
                return "Test"

        exporter = TestExporter()

        with pytest.raises(
            ValidationError, match="Cannot export schedule with no units"
        ):
            exporter.validate_schedule(empty_schedule)


# ========== CSV Exporter Tests ==========


class TestCSVExporter:
    """Tests for CSVExporter."""

    def test_export_success(self, sample_schedule, temp_output_dir):
        """Test successful CSV export."""
        exporter = CSVExporter()
        output_path = temp_output_dir / "test_schedule.csv"

        result_path = exporter.export(sample_schedule, output_path)

        assert result_path == output_path
        assert output_path.exists()
        assert output_path.suffix == ".csv"

    def test_export_default_path(self, sample_schedule, temp_output_dir, monkeypatch):
        """Test CSV export with default path."""
        exporter = CSVExporter()
        monkeypatch.setattr(exporter.settings, "output_dir", temp_output_dir)

        result_path = exporter.export(sample_schedule)

        assert result_path.name == "schedule_2025_10.csv"
        assert result_path.exists()

    def test_export_content_structure(self, sample_schedule, temp_output_dir):
        """Test CSV export content structure."""
        exporter = CSVExporter()
        output_path = temp_output_dir / "test_schedule.csv"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        # Check header
        assert rows[0] == [
            "Подразделение",
            "Дата",
            "День недели",
            "Тип дежурства",
            "Время",
            "Примечания",
        ]

        # Check data rows (3 shifts total)
        assert len(rows) == 4  # header + 3 shifts

        # Check first data row contains expected values
        assert "Всеволожский дозор" in rows[1][0]
        assert "01.10.2025" in rows[1][1]
        assert "ПДН" in rows[1][3]  # DutyType.PDN value is Russian "ПДН"

    def test_export_multiple_units(self, sample_schedule, temp_output_dir):
        """Test CSV export with multiple units."""
        exporter = CSVExporter()
        output_path = temp_output_dir / "multi_unit.csv"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        # Verify both units are present
        unit_names = [row[0] for row in rows[1:]]
        assert any("Всеволожский дозор" in name for name in unit_names)
        assert any("Заневское ГП" in name for name in unit_names)

    def test_export_encoding(self, sample_schedule, temp_output_dir):
        """Test CSV export handles UTF-8 encoding correctly."""
        exporter = CSVExporter()
        output_path = temp_output_dir / "encoding_test.csv"

        exporter.export(sample_schedule, output_path)

        # Read and verify Russian characters
        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()
            assert "Подразделение" in content
            assert "дежурство" in content

    def test_export_empty_schedule_fails(self, empty_schedule, temp_output_dir):
        """Test CSV export fails for empty schedule."""
        exporter = CSVExporter()
        output_path = temp_output_dir / "empty.csv"

        with pytest.raises(ValidationError):
            exporter.export(empty_schedule, output_path)

    def test_export_creates_directory(self, sample_schedule, temp_output_dir):
        """Test CSV export creates output directory if it doesn't exist."""
        exporter = CSVExporter()
        nested_path = temp_output_dir / "nested" / "dir" / "test.csv"

        result_path = exporter.export(sample_schedule, nested_path)

        assert result_path.exists()
        assert result_path.parent.exists()

    def test_get_file_extension(self):
        """Test CSV exporter returns correct file extension."""
        exporter = CSVExporter()
        assert exporter.get_file_extension() == "csv"

    def test_get_format_name(self):
        """Test CSV exporter returns correct format name."""
        exporter = CSVExporter()
        assert exporter.get_format_name() == "CSV"


# ========== JSON Exporter Tests ==========


class TestJSONExporter:
    """Tests for JSONExporter."""

    def test_export_success(self, sample_schedule, temp_output_dir):
        """Test successful JSON export."""
        exporter = JSONExporter()
        output_path = temp_output_dir / "test_schedule.json"

        result_path = exporter.export(sample_schedule, output_path)

        assert result_path == output_path
        assert output_path.exists()
        assert output_path.suffix == ".json"

    def test_export_default_path(self, sample_schedule, temp_output_dir, monkeypatch):
        """Test JSON export with default path."""
        exporter = JSONExporter()
        monkeypatch.setattr(exporter.settings, "output_dir", temp_output_dir)

        result_path = exporter.export(sample_schedule)

        assert result_path.name == "schedule_2025_10.json"
        assert result_path.exists()

    def test_export_content_structure(self, sample_schedule, temp_output_dir):
        """Test JSON export content structure."""
        exporter = JSONExporter()
        output_path = temp_output_dir / "test_schedule.json"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Check structure
        assert "metadata" in data
        assert "schedule" in data

        # Check metadata
        assert data["metadata"]["month"] == "Октябрь"
        assert data["metadata"]["year"] == 2025

        # Check schedule
        assert len(data["schedule"]) == 2  # 2 units
        assert len(data["schedule"][0]["shifts"]) == 2  # 2 shifts in first unit
        assert len(data["schedule"][1]["shifts"]) == 1  # 1 shift in second unit

    def test_export_pretty_json(self, sample_schedule, temp_output_dir, monkeypatch):
        """Test JSON export with pretty formatting."""
        exporter = JSONExporter()
        monkeypatch.setattr(exporter.settings, "pretty_json", True)
        output_path = temp_output_dir / "pretty.json"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()
            # Pretty JSON should have indentation
            assert "\n  " in content

    def test_export_compact_json(self, sample_schedule, temp_output_dir, monkeypatch):
        """Test JSON export with compact formatting."""
        exporter = JSONExporter()
        monkeypatch.setattr(exporter.settings, "pretty_json", False)
        output_path = temp_output_dir / "compact.json"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()
            # Compact JSON should be a single line
            lines = content.strip().split("\n")
            assert len(lines) == 1

    def test_export_encoding(self, sample_schedule, temp_output_dir):
        """Test JSON export handles UTF-8 encoding correctly."""
        exporter = JSONExporter()
        output_path = temp_output_dir / "encoding_test.json"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            # Verify Russian characters are preserved
            assert "Октябрь" in data["metadata"]["month"]

    def test_export_empty_schedule_fails(self, empty_schedule, temp_output_dir):
        """Test JSON export fails for empty schedule."""
        exporter = JSONExporter()
        output_path = temp_output_dir / "empty.json"

        with pytest.raises(ValidationError):
            exporter.export(empty_schedule, output_path)

    def test_get_file_extension(self):
        """Test JSON exporter returns correct file extension."""
        exporter = JSONExporter()
        assert exporter.get_file_extension() == "json"

    def test_get_format_name(self):
        """Test JSON exporter returns correct format name."""
        exporter = JSONExporter()
        assert exporter.get_format_name() == "JSON"


# ========== Excel Exporter Tests ==========


class TestExcelExporter:
    """Tests for ExcelExporter."""

    def test_export_success(self, sample_schedule, temp_output_dir):
        """Test successful Excel export."""
        exporter = ExcelExporter()
        output_path = temp_output_dir / "test_schedule.xlsx"

        result_path = exporter.export(sample_schedule, output_path)

        assert result_path == output_path
        assert output_path.exists()
        assert output_path.suffix == ".xlsx"

    def test_export_default_path(self, sample_schedule, temp_output_dir, monkeypatch):
        """Test Excel export with default path."""
        exporter = ExcelExporter()
        monkeypatch.setattr(exporter.settings, "output_dir", temp_output_dir)

        result_path = exporter.export(sample_schedule)

        assert result_path.name == "schedule_2025_10.xlsx"
        assert result_path.exists()

    def test_export_content_structure(self, sample_schedule, temp_output_dir):
        """Test Excel export content structure."""
        exporter = ExcelExporter()
        output_path = temp_output_dir / "test_schedule.xlsx"

        exporter.export(sample_schedule, output_path)

        # Load and verify workbook
        wb = load_workbook(output_path)
        ws = wb.active

        # Check title
        assert "График дежурств ДНД" in ws["A1"].value

        # Check headers (row 3)
        assert ws["A3"].value == "Подразделение"
        assert ws["B3"].value == "Дата"
        assert ws["C3"].value == "День недели"

        # Check data (starts from row 4)
        assert "Всеволожский дозор" in ws["A4"].value
        assert ws["D4"].value == "ПДН"  # DutyType.PDN value is Russian "ПДН"

    def test_export_with_metadata(self, sample_schedule, temp_output_dir, monkeypatch):
        """Test Excel export includes metadata when enabled."""
        exporter = ExcelExporter()
        monkeypatch.setattr(exporter.settings, "include_metadata", True)
        output_path = temp_output_dir / "with_metadata.xlsx"

        exporter.export(sample_schedule, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Find metadata section (after the data)
        found_source = False
        found_signatory = False

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "Источник:" in str(cell.value):
                    found_source = True
                if cell.value and "Подписант:" in str(cell.value):
                    found_signatory = True

        assert found_source
        assert found_signatory

    def test_export_styling(self, sample_schedule, temp_output_dir):
        """Test Excel export applies correct styling."""
        exporter = ExcelExporter()
        output_path = temp_output_dir / "styled.xlsx"

        exporter.export(sample_schedule, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Check header styling (row 3)
        header_cell = ws["A3"]
        # Note: openpyxl may return color without alpha channel prefix
        assert header_cell.fill.start_color.rgb in ("FF366092", "00366092")
        assert header_cell.font.color.rgb in ("FFFFFFFF", "00FFFFFF")
        assert header_cell.font.bold is True

        # Check borders are applied
        data_cell = ws["A4"]
        assert data_cell.border.left.style == "thin"

    def test_export_empty_schedule_fails(self, empty_schedule, temp_output_dir):
        """Test Excel export fails for empty schedule."""
        exporter = ExcelExporter()
        output_path = temp_output_dir / "empty.xlsx"

        with pytest.raises(ValidationError):
            exporter.export(empty_schedule, output_path)

    def test_export_column_widths(self, sample_schedule, temp_output_dir):
        """Test Excel export sets correct column widths."""
        exporter = ExcelExporter()
        output_path = temp_output_dir / "widths.xlsx"

        exporter.export(sample_schedule, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Check column widths
        assert ws.column_dimensions["A"].width == 35  # Подразделение
        assert ws.column_dimensions["B"].width == 12  # Дата
        assert ws.column_dimensions["F"].width == 50  # Примечания

    def test_get_file_extension(self):
        """Test Excel exporter returns correct file extension."""
        exporter = ExcelExporter()
        assert exporter.get_file_extension() == "xlsx"

    def test_get_format_name(self):
        """Test Excel exporter returns correct format name."""
        exporter = ExcelExporter()
        assert exporter.get_format_name() == "Excel"


# ========== Markdown Exporter Tests ==========


class TestMarkdownExporter:
    """Tests for MarkdownExporter."""

    def test_export_success(self, sample_schedule, temp_output_dir):
        """Test successful Markdown export."""
        exporter = MarkdownExporter()
        output_path = temp_output_dir / "test_schedule.md"

        result_path = exporter.export(sample_schedule, output_path)

        assert result_path == output_path
        assert output_path.exists()
        assert output_path.suffix == ".md"

    def test_export_default_path(self, sample_schedule, temp_output_dir, monkeypatch):
        """Test Markdown export with default path."""
        exporter = MarkdownExporter()
        monkeypatch.setattr(exporter.settings, "output_dir", temp_output_dir)

        result_path = exporter.export(sample_schedule)

        assert result_path.name == "schedule_2025_10.md"
        assert result_path.exists()

    def test_export_content_structure(self, sample_schedule, temp_output_dir):
        """Test Markdown export content structure."""
        exporter = MarkdownExporter()
        output_path = temp_output_dir / "test_schedule.md"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Check main sections
        assert "# График дежурств ДНД" in content
        assert "## Статистика" in content
        assert "## График дежурств" in content

        # Check table structure
        assert "| Подразделение |" in content
        assert "|---------------|" in content

        # Check data
        assert "Всеволожский дозор" in content
        assert "01.10.2025" in content
        assert "ПДН" in content  # DutyType.PDN value is Russian "ПДН"

    def test_export_with_metadata(self, sample_schedule, temp_output_dir, monkeypatch):
        """Test Markdown export includes metadata when enabled."""
        exporter = MarkdownExporter()
        monkeypatch.setattr(exporter.settings, "include_metadata", True)
        output_path = temp_output_dir / "with_metadata.md"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        assert "## Информация о документе" in content
        assert "**Источник:**" in content
        assert "**Подписант:**" in content

    def test_export_statistics(self, sample_schedule, temp_output_dir):
        """Test Markdown export includes statistics."""
        exporter = MarkdownExporter()
        output_path = temp_output_dir / "stats.md"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Check statistics section
        assert "**Всего подразделений:** 2" in content
        assert "**Всего дежурств:** 3" in content
        assert "**По типам:**" in content

    def test_export_empty_schedule_fails(self, empty_schedule, temp_output_dir):
        """Test Markdown export fails for empty schedule."""
        exporter = MarkdownExporter()
        output_path = temp_output_dir / "empty.md"

        with pytest.raises(ValidationError):
            exporter.export(empty_schedule, output_path)

    def test_export_footer(self, sample_schedule, temp_output_dir, monkeypatch):
        """Test Markdown export includes footer with version."""
        exporter = MarkdownExporter()
        monkeypatch.setattr(exporter.settings, "app_version", "1.0.0")
        output_path = temp_output_dir / "footer.md"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        assert "Schedule DND v1.0.0" in content

    def test_get_file_extension(self):
        """Test Markdown exporter returns correct file extension."""
        exporter = MarkdownExporter()
        assert exporter.get_file_extension() == "md"

    def test_get_format_name(self):
        """Test Markdown exporter returns correct format name."""
        exporter = MarkdownExporter()
        assert exporter.get_format_name() == "Markdown"


# ========== HTML Exporter Tests ==========


class TestHTMLExporter:
    """Tests for HTMLExporter."""

    def test_export_success(self, sample_schedule, temp_output_dir):
        """Test successful HTML export."""
        exporter = HTMLExporter()
        output_path = temp_output_dir / "test_schedule.html"

        result_path = exporter.export(sample_schedule, output_path)

        assert result_path == output_path
        assert output_path.exists()
        assert output_path.suffix == ".html"

    def test_export_default_path(self, sample_schedule, temp_output_dir, monkeypatch):
        """Test HTML export with default path."""
        exporter = HTMLExporter()
        monkeypatch.setattr(exporter.settings, "output_dir", temp_output_dir)

        result_path = exporter.export(sample_schedule)

        assert result_path.name == "schedule_2025_10.html"
        assert result_path.exists()

    def test_export_content_structure(self, sample_schedule, temp_output_dir):
        """Test HTML export content structure."""
        exporter = HTMLExporter()
        output_path = temp_output_dir / "test_schedule.html"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Check HTML structure
        assert "<!DOCTYPE html>" in content
        assert '<html lang="ru">' in content
        assert "<head>" in content
        assert "<body>" in content

        # Check title
        assert "График дежурств ДНД" in content

        # Check table
        assert "<table>" in content
        assert "<thead>" in content
        assert "<tbody>" in content

        # Check data
        assert "Всеволожский дозор" in content
        assert "01.10.2025" in content

    def test_export_with_metadata(self, sample_schedule, temp_output_dir, monkeypatch):
        """Test HTML export includes metadata when enabled."""
        exporter = HTMLExporter()
        monkeypatch.setattr(exporter.settings, "include_metadata", True)
        output_path = temp_output_dir / "with_metadata.html"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        assert "metadata" in content.lower()
        assert "Создано:" in content
        assert "Источник:" in content

    def test_export_styling(self, sample_schedule, temp_output_dir):
        """Test HTML export includes CSS styling."""
        exporter = HTMLExporter()
        output_path = temp_output_dir / "styled.html"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Check CSS is present
        assert "<style>" in content
        assert ".container" in content
        assert ".header" in content
        assert "table" in content

        # Check responsive design
        assert "@media print" in content

    def test_export_statistics_cards(self, sample_schedule, temp_output_dir):
        """Test HTML export includes statistics cards."""
        exporter = HTMLExporter()
        output_path = temp_output_dir / "stats.html"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Check statistics section
        assert "stats" in content
        assert "stat-card" in content

    def test_export_empty_schedule_fails(self, empty_schedule, temp_output_dir):
        """Test HTML export fails for empty schedule."""
        exporter = HTMLExporter()
        output_path = temp_output_dir / "empty.html"

        with pytest.raises(ValidationError):
            exporter.export(empty_schedule, output_path)

    def test_export_encoding(self, sample_schedule, temp_output_dir):
        """Test HTML export handles UTF-8 encoding correctly."""
        exporter = HTMLExporter()
        output_path = temp_output_dir / "encoding_test.html"

        exporter.export(sample_schedule, output_path)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Check meta charset
        assert '<meta charset="UTF-8">' in content

        # Verify Russian characters
        assert "Подразделение" in content

    def test_get_file_extension(self):
        """Test HTML exporter returns correct file extension."""
        exporter = HTMLExporter()
        assert exporter.get_file_extension() == "html"

    def test_get_format_name(self):
        """Test HTML exporter returns correct format name."""
        exporter = HTMLExporter()
        assert exporter.get_format_name() == "HTML"


# ========== Exporter Factory Tests ==========


class TestExporterFactory:
    """Tests for ExporterFactory."""

    def test_create_csv_exporter(self):
        """Test factory creates CSV exporter."""
        exporter = ExporterFactory.create(ExportFormat.CSV)
        assert isinstance(exporter, CSVExporter)

    def test_create_json_exporter(self):
        """Test factory creates JSON exporter."""
        exporter = ExporterFactory.create(ExportFormat.JSON)
        assert isinstance(exporter, JSONExporter)

    def test_create_excel_exporter(self):
        """Test factory creates Excel exporter."""
        exporter = ExporterFactory.create(ExportFormat.EXCEL)
        assert isinstance(exporter, ExcelExporter)

    def test_create_markdown_exporter(self):
        """Test factory creates Markdown exporter."""
        exporter = ExporterFactory.create(ExportFormat.MARKDOWN)
        assert isinstance(exporter, MarkdownExporter)

    def test_create_html_exporter(self):
        """Test factory creates HTML exporter."""
        exporter = ExporterFactory.create(ExportFormat.HTML)
        assert isinstance(exporter, HTMLExporter)

    def test_unsupported_format_raises_error(self):
        """Test factory raises error for unsupported format."""
        # Test with ExportFormat.from_string which should raise ValueError for invalid format
        with pytest.raises(ValueError, match="Invalid export format"):
            invalid_format = ExportFormat.from_string("unsupported")

    def test_get_supported_formats(self):
        """Test factory returns list of supported formats."""
        formats = ExporterFactory.get_supported_formats()

        assert ExportFormat.CSV in formats
        assert ExportFormat.JSON in formats
        assert ExportFormat.EXCEL in formats
        assert ExportFormat.MARKDOWN in formats
        assert ExportFormat.HTML in formats
        assert len(formats) == 5

    def test_is_supported(self):
        """Test factory can check if format is supported."""
        assert ExporterFactory.is_supported(ExportFormat.CSV) is True
        assert ExporterFactory.is_supported(ExportFormat.JSON) is True
        assert ExporterFactory.is_supported(ExportFormat.EXCEL) is True


# ========== Integration Tests ==========


class TestExporterIntegration:
    """Integration tests for exporters working together."""

    def test_all_exporters_produce_valid_files(self, sample_schedule, temp_output_dir):
        """Test all exporters can export the same schedule."""
        formats = [
            (ExportFormat.CSV, "csv"),
            (ExportFormat.JSON, "json"),
            (ExportFormat.EXCEL, "xlsx"),
            (ExportFormat.MARKDOWN, "md"),
            (ExportFormat.HTML, "html"),
        ]

        for format_type, ext in formats:
            exporter = ExporterFactory.create(format_type)
            output_path = temp_output_dir / f"schedule.{ext}"

            result_path = exporter.export(sample_schedule, output_path)

            assert result_path.exists()
            assert result_path.stat().st_size > 0

    def test_minimal_schedule_exports_in_all_formats(
        self, minimal_schedule, temp_output_dir
    ):
        """Test minimal schedule can be exported in all formats."""
        formats = [
            (ExportFormat.CSV, "csv"),
            (ExportFormat.JSON, "json"),
            (ExportFormat.EXCEL, "xlsx"),
            (ExportFormat.MARKDOWN, "md"),
            (ExportFormat.HTML, "html"),
        ]

        for format_type, ext in formats:
            exporter = ExporterFactory.create(format_type)
            output_path = temp_output_dir / f"minimal.{ext}"

            result_path = exporter.export(minimal_schedule, output_path)

            assert result_path.exists()

    def test_consistent_filename_across_exporters(self, sample_schedule):
        """Test all exporters generate consistent base filenames."""
        formats = [
            ExportFormat.CSV,
            ExportFormat.JSON,
            ExportFormat.EXCEL,
            ExportFormat.MARKDOWN,
            ExportFormat.HTML,
        ]
        filenames = []

        for format_type in formats:
            exporter = ExporterFactory.create(format_type)
            filename = exporter.get_default_filename(sample_schedule)
            base_name = filename.rsplit(".", 1)[0]
            filenames.append(base_name)

        # All base filenames should be the same
        assert len(set(filenames)) == 1
        assert filenames[0] == "schedule_2025_10"
